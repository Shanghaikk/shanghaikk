#!/usr/bin/env python3
"""
DeepSeek 余额本地备份脚本
每10分钟运行一次（launchd），记录时间取整到 :00/:10/:20/:30/:40/:50
自动计算与上一条的差额，输出到 Excel + JSON，并推到 GitHub
"""

import os
import sys
import math
import json
import subprocess
from datetime import datetime, timezone, timedelta
import urllib.request

# ─── 配置 ────────────────────────────────────────────────
API_KEY = "sk-9e4d05ee82f34cd0a8240eeffee998a9"
API_URL = "https://api.deepseek.com/user/balance"
DATA_DIR = os.path.dirname(os.path.abspath(__file__))
GIT_REPO = os.path.join(DATA_DIR, "repo")
GIT_FILE = "deepseek/balance-data.json"
GIT_BRANCH = "main"
EXCEL_PATH = os.path.join(DATA_DIR, "deepseek_balance.xlsx")
TZ = timezone(timedelta(hours=8))
# ─────────────────────────────────────────────────────────


def git(*args, capture=True):
    cmd = ["git", "-C", GIT_REPO] + list(args)
    env = os.environ.copy()
    env["GIT_SSH_COMMAND"] = "ssh -i ~/.ssh/id_ed25519_github"
    kwargs = {"env": env}
    if capture:
        kwargs["capture_output"] = True
        kwargs["text"] = True
    r = subprocess.run(cmd, **kwargs)
    return r


def fetch_balance():
    req = urllib.request.Request(
        API_URL,
        headers={"Authorization": f"Bearer {API_KEY}", "Accept": "application/json"},
    )
    with urllib.request.urlopen(req, timeout=15) as resp:
        return json.loads(resp.read().decode())


def round_to_10min(dt):
    """四舍五入到最近的10分钟，返回格式 '2026-07-22 11:50'"""
    ts = dt.timestamp()
    rounded_ts = round(ts / 600) * 600
    rounded = datetime.fromtimestamp(rounded_ts, tz=TZ)
    return rounded.strftime("%Y-%m-%d %H:%M")


def load_history():
    path = os.path.join(GIT_REPO, GIT_FILE)
    if os.path.exists(path):
        try:
            with open(path) as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            pass
    return []


def save_history(history):
    path = os.path.join(GIT_REPO, GIT_FILE)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def push_to_github(history_count):
    try:
        git("pull", "origin", GIT_BRANCH)
        git("add", GIT_FILE)
        r = git("commit", "-m", f"🤖 balance-data: {history_count} records")
        if r.returncode != 0:
            return True
        r = git("push", "origin", GIT_BRANCH)
        if r.returncode != 0:
            print(f"[WARN] git push: {r.stderr.strip()}")
            return False
        print(f"[GIT] pushed {history_count} records")
        return True
    except Exception as e:
        print(f"[ERR] git: {e}")
        return False


def rebuild_excel(history):
    """根据 JSON 历史完整重建 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "余额记录"

    headers = ["时间", "总余额", "变化", "赠送余额", "充值余额", "币种", "状态"]
    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 10
    ws.freeze_panes = "A2"

    for h in history:
        t = h.get("t", "")
        b = h.get("b", "")
        d = h.get("d", "")
        g = h.get("g", "")
        tp = h.get("tp", "")
        cur = h.get("c", "CNY")
        st = h.get("s", "正常")
        row = [t, b, d, g, tp, cur, st]
        ws.append(row)

    wb.save(EXCEL_PATH)


def save_record_to_excel(entry):
    """追加单条记录到 Excel（增量）"""
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "余额记录"
        headers = ["时间", "总余额", "变化", "赠送余额", "充值余额", "币种", "状态"]
        hfont = Font(bold=True, color="FFFFFF")
        hfill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 10
        ws.freeze_panes = "A2"

    ws.append([
        entry.get("t", ""),
        entry.get("b", ""),
        entry.get("d", ""),
        entry.get("g", ""),
        entry.get("tp", ""),
        entry.get("c", "CNY"),
        entry.get("s", "正常"),
    ])
    wb.save(EXCEL_PATH)


CONFIG_DIR = os.path.dirname(os.path.abspath(__file__))


def update_usage_stats():
    """同步更新用量统计（请求次数/tokens/消费）"""
    track_script = os.path.join(CONFIG_DIR, "track_usage.py")
    if os.path.exists(track_script):
        subprocess.run([sys.executable, track_script], capture_output=True, timeout=30)


def main():
    try:
        data = fetch_balance()
        is_available = data.get("is_available", False)
        info = data.get("balance_infos", [{}])[0]
        bal = float(info.get("total_balance", 0))
        granted = float(info.get("granted_balance", 0))
        topped = float(info.get("topped_up_balance", 0))
        currency = info.get("currency", "CNY")

        # 时间取整到10分钟
        time_key = round_to_10min(datetime.now(TZ))

        # ── 加载历史 ──
        history = load_history()

        # ── 去重：这个10分钟槽已经记录过了就不再记 ──
        existing_keys = {h["t"] for h in history if isinstance(h.get("b"), (int, float))}
        if time_key in existing_keys:
            print(f"[SKIP] {time_key} already recorded")
            # 仍然 push 一下确保同步
            if len(history) >= 1:
                push_to_github(len(history))
            return

        # ── 计算差值 ──
        prev_bal = None
        for h in reversed(history):
            if isinstance(h.get("b"), (int, float)):
                prev_bal = h["b"]
                break
        delta = round(bal - prev_bal, 2) if prev_bal is not None else 0.00

        # ── 构建记录 ──
        entry = {
            "t": time_key,
            "b": round(bal, 2),
            "d": delta,
            "g": round(granted, 4),
            "tp": round(topped, 4),
            "c": currency,
            "s": "正常" if is_available else "已停用",
        }
        history.append(entry)

        # 限制最大条数（约 347 天）
        if len(history) > 50000:
            history = history[-50000:]

        save_history(history)
        print(f"[JSON] {time_key} | ¥{bal:.2f} | Δ{delta:+.2f}")

        # ── Excel ──
        save_record_to_excel(entry)
        print(f"[EXCEL] {time_key} | ¥{bal:.2f} | Δ{delta:+.2f}")

        # ── 用量统计 ──
    update_usage_stats()

    # ── GitHub ──
        push_to_github(len(history))

    except Exception as e:
        print(f"[ERR] {datetime.now(TZ).isoformat()} | {e}", file=sys.stderr)
        raise


def batch_push():
    history = load_history()
    if history:
        push_to_github(len(history))


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--push":
        batch_push()
    else:
        main()
